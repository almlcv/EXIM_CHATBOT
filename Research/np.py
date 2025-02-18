from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
import pandas as pd
import uvicorn
import httpx
import asyncio
from typing import Dict, List, Optional
from datetime import datetime
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from functools import lru_cache

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constants
EXCEL_FILE = "Namdeo.xlsx"
API_URL = "http://43.205.59.159:9000/api/download-report/24-25/Pending"
REFRESH_INTERVAL = 300  # 5 minutes
COLUMN_WIDTHS = {
    'JOB NO AND DATE': 40,
    'SUPPLIER/ EXPORTER': 40,
    'INVOICE NUMBER AND DATE': 25,
    'INVOICE VALUE AND UNIT PRICE': 35,
    'BL NUMBER AND DATE': 25,
    'COMMODITY': 50,
    'NET WEIGHT': 15,
    'PORT': 25,
    'ARRIVAL DATE': 15,
    'FREE TIME': 12,
    'DETENTION FROM': 15,
    'SHIPPING LINE': 40,
    'CONTAINER NUM & SIZE': 30,
    'WEIGHT EXCESS/SHORTAGE': 20,
    'NUMBER OF CONTAINERS': 15,
    'BE NUMBER AND DATE': 35,
    'REMARKS': 60,
    'DETAILED STATUS': 35
}

HEADERS = [
    "JOB NO AND DATE", "IMPORTER", "SUPPLIER/ EXPORTER", "INVOICE NUMBER AND DATE",
    "INVOICE VALUE AND UNIT PRICE", "BL NUMBER AND DATE", "COMMODITY", "NET WEIGHT",
    "PORT", "ARRIVAL DATE", "FREE TIME", "DETENTION FROM", "SHIPPING LINE",
    "CONTAINER NUM & SIZE", "NUMBER OF CONTAINERS", "BE NUMBER AND DATE", "REMARKS", 
    "DETAILED STATUS"
]

app = FastAPI(title="Container Details API")

class ExcelFormatter:
    @staticmethod
    def style_header(ws):
        """Apply styling to the header row"""
        header_style = {
            'fill': PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
            'font': Font(bold=True),
            'alignment': Alignment(horizontal="center", vertical="center")
        }
        
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_style['fill']
            cell.font = header_style['font']
            cell.alignment = header_style['alignment']
            ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS.get(header, 20)

    @staticmethod
    def style_data(ws, row_count: int):
        """Apply styling to data cells"""
        alignment = Alignment(horizontal="center", vertical="center")
        for row in range(2, row_count + 1):
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).alignment = alignment

class DataProcessor:
    @staticmethod
    async def fetch_api_data():
        """Fetch data from API asynchronously"""
        async with httpx.AsyncClient() as client:
            try:
                response = await client.get(API_URL)
                response.raise_for_status()
                return response.json().get('data', [])
            except Exception as e:
                logger.error(f"API fetch error: {e}")
                return []

    @staticmethod
    def format_row_data(row: Dict) -> List:
        """Format a single row of data"""
        containers = row.get('container_nos', [])
        return [
            f"{row.get('job_no', '')} | {row.get('job_date', '')} | {row.get('custom_house', '')} | {row.get('type_of_b_e', '')}",
            row.get('importer', ''),
            row.get('supplier_exporter', ''),
            f"{row.get('invoice_number', '')} | {row.get('invoice_date', '')}",
            f"{row.get('inv_currency', '')} {row.get('invoice_value', '')} | {row.get('unit_price', '')}",
            f"{row.get('awb_bl_no', '')} | {row.get('awb_bl_date', '')}",
            row.get('description', ''),
            row.get('job_net_weight', ''),
            f"POL: {row.get('loading_port', '')} POD: {row.get('port_of_reporting', '')}",
            ",\n".join(c.get('arrival_date', '') for c in containers),
            row.get('free_time', ''),
            ",\n".join(c.get('detention_from', '') for c in containers),
            row.get('shipping_line_airline', ''),
            ", ".join(f"{c.get('container_number', '')} - {c.get('size', '')}" for c in containers),
            row.get('no_of_container', ''),
            f"{row.get('be_no', '')} | {row.get('be_date', '')}",
            (f"Discharge_Date: {row.get('discharge_date', '')} | "
             f"Arrival_Date: {row.get('assessment_date', '')} | "
             f"Duty_Paid_Date: {row.get('duty_paid_date', '')} | "
             f"DO_Validity_Upto_Job_Level: {row.get('do_validity_upto_job_level', '')}"),
            row.get('detailed_status', '')
        ]

class ContainerService:
    @staticmethod
    @lru_cache(maxsize=100)
    def get_container_details(container_number: str) -> str:
        """Get container details with caching"""
        try:
            df = pd.read_excel(EXCEL_FILE, dtype=str)
            df = df.fillna("Not Available")
            
            container_number = container_number.upper()
            filtered_df = df[df["CONTAINER NUM & SIZE"].str.upper().str.contains(container_number)]
            
            if filtered_df.empty:
                return f"\n❌ No data found for container number: {container_number}\n"
            
            row = filtered_df.iloc[0]
            details = {col: row.get(col, "Not Available") for col in HEADERS}
            
            return ContainerService._format_container_output(container_number, details)
        except Exception as e:
            logger.error(f"Error getting container details: {e}")
            raise HTTPException(status_code=500, detail="Error processing container details")

    @staticmethod
    def _format_container_output(container_number: str, details: Dict) -> str:
        """Format container details output"""
        return "\n".join([
            f'📦 **Container Details for: {container_number}**'] + 
            [f'✅ **{header}:** {details[header]}' for header in HEADERS]
        )

async def update_excel_file():
    """Update Excel file periodically"""
    while True:
        try:
            data = await DataProcessor.fetch_api_data()
            if data:
                wb = Workbook()
                ws = wb.active
                ws.append(HEADERS)
                
                for row in data:
                    ws.append(DataProcessor.format_row_data(row))
                
                formatter = ExcelFormatter()
                formatter.style_header(ws)
                formatter.style_data(ws, len(data) + 1)
                
                wb.save(EXCEL_FILE)
                logger.info(f"Excel file updated at {datetime.now()}")
                ContainerService.get_container_details.cache_clear()
        except Exception as e:
            logger.error(f"Error updating Excel file: {e}")
        
        await asyncio.sleep(REFRESH_INTERVAL)

@app.on_event("startup")
async def startup_event():
    """Start background tasks"""
    asyncio.create_task(update_excel_file())

@app.get("/container/{container_number}")
async def get_container_details(container_number: str):
    """API endpoint to get container details"""
    return ContainerService.get_container_details(container_number)

@app.get("/download-excel")
async def download_excel():
    """API endpoint to download the Excel file"""
    if not Path(EXCEL_FILE).exists():
        raise HTTPException(status_code=404, detail="Excel file not found")
    return FileResponse(path=EXCEL_FILE, filename=EXCEL_FILE, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)