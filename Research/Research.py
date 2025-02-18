import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal
import json
import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def validate_data(data):
    """Validate the API response data"""
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON string: {e}")
            raise ValueError("API returned invalid JSON string")
    
    # Handle the nested data structure
    if isinstance(data, dict) and 'data' in data:
        logger.info("Found nested data structure, extracting 'data' field")
        data = data['data']
    
    if not isinstance(data, list):
        logger.error(f"Expected list but got {type(data)}")
        raise ValueError("API response is not in expected format (list)")
    
    return data

def format_date(date_str):
    """Format date string to dd/mm/yyyy format"""
    if not date_str:
        return ""
    try:
        # Try multiple date formats
        for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y']:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%d/%m/%Y')
            except ValueError:
                continue
        return date_str
    except Exception as e:
        logger.warning(f"Failed to parse date {date_str}: {e}")
        return date_str

def format_container_dates(containers, date_field):
    """Format container dates and handle multiple dates"""
    if not containers:
        return ""
    
    valid_dates = []
    for container in containers:
        date = container.get(date_field)
        if date:
            formatted_date = format_date(date)
            if formatted_date:
                valid_dates.append(formatted_date)

    if not valid_dates:
        return ""
   
    if all(date == valid_dates[0] for date in valid_dates):
        return valid_dates[0]
    return ",\n".join(valid_dates)

def get_cell_color(detailed_status):
    """Get cell color based on detailed status"""
    color_map = {
        "ETA Date Pending": "FFFFFF",  # white
        "Estimated Time of Arrival": "FFFF99",  # Light Yellow
        "Custom Clearance Completed": "CCFFFF",  # Light Blue
        "PCV Done, Duty Payment Pending": "FFDBFF",  # Light Blue
        "Discharged": "FFCC99",  # Light Orange
        "BE Noted, Arrival Pending": "99CCFF",  # Light Purple
        "BE Noted, Clearance Pending": "99CCFF",  # Light Purple
        "Gateway IGM Filed": "FFCC99",  # Light Orange
    }
    return color_map.get(detailed_status, "FFFFFF")

def convert_to_excel(data, output_filename):
    """Convert JSON data to Excel report"""
    try:
        # Validate and preprocess data
        data = validate_data(data)

        # Filter rows without bill_no
        rows = [row for row in data if isinstance(row, dict) and not row.get('bill_no')]
        
        if not rows:
            logger.warning("No valid data rows to export")
            return
        
        logger.info(f"Processing {len(rows)} rows of data")
        
        wb = Workbook()
        ws = wb.active
        
        # Get unique statuses for reference row
        unique_statuses = list(set(row.get('detailed_status', '') 
                                 for row in rows 
                                 if row.get('detailed_status')))
        
        date_of_report = datetime.now().strftime('%d-%b-%Y %I:%M:%S %p')
        
        # Define headers
        headers = [
            "JOB NO AND DATE",
            "SUPPLIER/ EXPORTER",
            "INVOICE NUMBER AND DATE",
            "INVOICE VALUE AND UNIT PRICE",
            "BL NUMBER AND DATE",
            "COMMODITY",
            "NET WEIGHT",
            "PORT",
            "ARRIVAL DATE",
            "FREE TIME",
            "DETENTION FROM",
            "SHIPPING LINE",
            "CONTAINER NUM & SIZE",
            "WEIGHT EXCESS/SHORTAGE",
            "NUMBER OF CONTAINERS",
            "BE NUMBER AND DATE",
            "REMARKS",
            "DETAILED STATUS"
        ]

        # Add reference row
        reference_row = ["REFERENCE"] + unique_statuses
        ws.append(reference_row)
        for idx, cell in enumerate(ws[1], 1):
            cell.fill = PatternFill(start_color="4472C4" if idx == 1 else get_cell_color(cell.value),
                                fill_type="solid")
            cell.font = Font(color="FFFFFF" if idx == 1 else "000000", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add title row
        ws.merge_cells(f'A3:{get_column_letter(len(headers))}3')
        title_cell = ws['A3']
        title_cell.value = f"{rows[0].get('importer', '')}: Status as of {date_of_report}"
        title_cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
        title_cell.font = Font(color="FFFFFF", size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add headers row
        ws.append(headers)
        for cell in ws[4]:
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row in rows:
            try:
                # Format job number and date
                job_no_date = f"{row.get('job_no', '')} | {format_date(row.get('job_date', ''))} | {row.get('custom_house', '')} | {row.get('type_of_b_e', '')}"
                
                # Format invoice details
                invoice_details = f"{row.get('invoice_number', '')} | {format_date(row.get('invoice_date', ''))}"
                
                # Calculate invoice value safely
                try:
                    cif_amount = Decimal(str(row.get('cif_amount', 0)))
                    exrate = Decimal(str(row.get('exrate', 1)))
                    inv_value = (cif_amount / exrate).quantize(Decimal('0.01'))
                except (decimal.InvalidOperation, TypeError):
                    inv_value = Decimal('0.00')
                
                # Format container information
                containers = row.get('container_nos', [])
                container_numbers = ",\n".join(
                    f"{c.get('container_number', '')} - {c.get('size', '')}" 
                    for c in containers
                )
                
                data_row = [
                    job_no_date,
                    row.get('supplier_exporter', ''),
                    invoice_details,
                    f"{row.get('inv_currency', '')} | {inv_value} | {row.get('unit_price', '')}",
                    f"{row.get('awb_bl_no', '')} | {format_date(row.get('awb_bl_date', ''))}",
                    row.get('description', ''),
                    row.get('job_net_weight', ''),
                    f"POL: {row.get('loading_port', '').split('(')[0]}\nPOD: {row.get('port_of_reporting', '').split('(')[0]}",
                    format_container_dates(containers, 'arrival_date'),
                    row.get('free_time', ''),
                    format_container_dates(containers, 'detention_from'),
                    row.get('shipping_line_airline', ''),
                    container_numbers,
                    ",\n".join(str(c.get('weight_shortage', '')) for c in containers),
                    row.get('no_of_container', '')[:-2] if row.get('no_of_container') else '',
                    f"{row.get('be_no', '')} | {format_date(row.get('be_date', ''))}",
                    format_remarks(row),
                    row.get('detailed_status', '')
                ]
                
                ws.append(data_row)
                
                # Format data row
                row_num = ws.max_row
                for cell in ws[row_num]:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            except Exception as e:
                logger.error(f"Error processing row: {e}", exc_info=True)
                continue

        add_summary_section(ws, rows)
        set_column_widths(ws, headers)
        
        wb.save(output_filename)
        logger.info(f"Successfully saved report to {output_filename}")
        
    except Exception as e:
        logger.error(f"Error in convert_to_excel: {e}", exc_info=True)
        raise

def format_remarks(row):
    """Format remarks section with all dates and additional information"""
    remarks = []
    
    if row.get('discharge_date'):
        remarks.append(f"Discharge Date: {format_date(row['discharge_date'])}")
    elif row.get('vessel_berthing'):
        remarks.append(f"ETA: {format_date(row['vessel_berthing'])}")
        
    date_fields = {
        'assessment_date': 'Assessment Date',
        'rail_out_date': 'Rail-Out',
        'examination_date': 'Examination Date',
        'duty_paid_date': 'Duty Paid Date',
        'out_of_charge': 'OOC Date',
        'sims_reg_no': 'SIMS Reg No',
        'sims_date': 'SIMS Reg Date',
        'pims_reg_no': 'PIMS Reg No',
        'pims_date': 'PIMS Reg Date',
        'nfmims_reg_no': 'NFMIMS Reg No',
        'nfmims_date': 'NFMIMS Reg Date'
    }
    
    for field, label in date_fields.items():
        if row.get(field):
            remarks.append(f"{label}: {format_date(row[field])}")
            
    if row.get('obl_telex_bl'):
        doc_type = "ORG-RCVD" if row['obl_telex_bl'] == "OBL" else "DOC-RCVD"
        remarks.append(f"{doc_type}: {row.get('document_received_date', '')}")
        
    if row.get('do_validity'):
        remarks.append(f"DO VALIDITY: {format_date(row['do_validity'])}")
        
    if row.get('remarks'):
        remarks.append(f"Remarks: {row['remarks']}")
        
    return " | ".join(remarks)

def add_summary_section(ws, rows):
    """Add summary section at the bottom of the worksheet"""
    # Add blank rows
    ws.append([])
    ws.append([])
    
    summary_row = ws.max_row + 1
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "SUMMARY"
    summary_cell.fill = PatternFill(start_color="92D050", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    container_counts = {
        '20_arrived': 0, '40_arrived': 0,
        '20_transit': 0, '40_transit': 0
    }
    
    for row in rows:
        for container in row.get('container_nos', []):
            size = container.get('size', '')
            is_arrived = bool(container.get('arrival_date'))
            key = f"{size}_{'arrived' if is_arrived else 'transit'}"
            if key in container_counts:
                container_counts[key] += 1
    
    ws.append(['ARRIVED', '', 'IN TRANSIT', '', 'TOTAL'])
    ws.append([
        container_counts['20_arrived'],
        container_counts['40_arrived'],
        container_counts['20_transit'],
        container_counts['40_transit'],
        sum(container_counts.values())
    ])

def set_column_widths(ws, headers):
    """Set column widths based on content"""
    column_widths = {
        'JOB NO AND DATE': 25,
        'SUPPLIER/ EXPORTER': 30,
        'INVOICE NUMBER AND DATE': 25,
        'INVOICE VALUE AND UNIT PRICE': 35,
        'BL NUMBER AND DATE': 25,
        'COMMODITY': 50,
        'NET WEIGHT': 15,
        'PORT': 25,
        'ARRIVAL DATE': 15,
        'FREE TIME': 12,
        'DETENTION FROM': 15,
        'SHIPPING LINE': 40,
        'CONTAINER NUM & SIZE': 30,
        'WEIGHT EXCESS/SHORTAGE': 20,
        'NUMBER OF CONTAINERS': 15,
        'BE NUMBER AND DATE': 25,
        'REMARKS': 45,
        'DETAILED STATUS': 25
    }

    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = column_widths.get(header, 15)

def main():
    """Main function to fetch data and generate report"""
    api_url = "http://43.205.59.159:9000/api/24-25/jobs/Pending/all"
    
    try:
        logger.info(f"Fetching data from {api_url}")
        response = requests.get(api_url)
        response.raise_for_status()  # Raise exception for bad status codes
        
        # Log response details for debugging
        logger.debug(f"Response status code: {response.status_code}")
        logger.debug(f"Response content type: {response.headers.get('content-type', 'unknown')}")
        logger.debug(f"First 200 characters of response: {response.text[:200]}")
        
        try:
            raw_data = response.json()
            logger.debug(f"Successfully parsed JSON response")
            if isinstance(raw_data, dict) and 'data' in raw_data:
                data = raw_data['data']
            else:
                data = raw_data
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON response: {e}")
            raise ValueError("Invalid JSON received from API")
        
        output_filename = f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        convert_to_excel(data, output_filename)
        logger.info(f"Excel report generated: {output_filename}")
    
    except requests.RequestException as e:
        logger.error(f"API request failed: {e}", exc_info=True)
        raise

if __name__ == "__main__":
    main()