

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from datetime import datetime
logger = logging.getLogger(__name__)
from fastapi import FastAPI, BackgroundTasks
import requests
import asyncio
from datetime import datetime
import logging
from openpyxl import Workbook
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from datetime import datetime


app = FastAPI()

# API URL
API_URL = "http://43.205.59.159:9000/api/download-report/24-25/Pending"

# Store last fetch status
fetch_status = {"last_run": None, "last_report": None, "error": None}


def validate_data(data):
    """Validate and extract the required API response data"""
    try:
        if isinstance(data, str):
            data = json.loads(data)

        if isinstance(data, dict) and "data" in data:
            data = data["data"]

        if not isinstance(data, list):
            raise ValueError("API response is not in expected format (list)")
        
        return data
    except Exception as e:
        logger.error(f"Data validation failed: {e}")
        raise


def format_date(date_str):
    """Formats the date field to avoid empty values"""
    return date_str if date_str else ""

def format_container_dates(containers, key):
    """Formats the container-related dates"""
    return ",\n".join(c.get(key, '') for c in containers)

def format_remarks(row):
    """Formats remarks field"""
    return row.get('remarks', '') if row.get('remarks') else ''



def set_column_widths(ws, headers):
    """Set column widths based on predefined values."""
    column_widths = {
        'JOB NO AND DATE': 40,
        'SUPPLIER/ EXPORTER': 40,
        'INVOICE NUMBER AND DATE': 25,
        'INVOICE VALUE AND UNIT PRICE': 35,
        'BL NUMBER AND DATE': 25,
        'COMMODITY': 50,
        'NET WEIGHT': 15,
        'PORT': 25,
        'ARRIVAL DATE': 15,
        'FREE TIME': 12,
        'DETENTION FROM': 15,
        'SHIPPING LINE': 40,
        'CONTAINER NUM & SIZE': 30,
        'WEIGHT EXCESS/SHORTAGE': 20,
        'NUMBER OF CONTAINERS': 15,
        'BE NUMBER AND DATE': 35,
        'REMARKS': 60,
        'DETAILED STATUS': 35
    }

    for col_num, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = column_widths.get(header, 20)  # Default width 20

def style_header(ws, headers):
    """Apply styling to the header row (bold, centered, highlighted)."""
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light Yellow
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

def style_data(ws, start_row, end_row, num_columns):
    """Apply center alignment to all data cells."""
    data_alignment = Alignment(horizontal="center", vertical="center")

    for row in range(start_row, end_row + 1):
        for col in range(1, num_columns + 1):
            ws.cell(row=row, column=col).alignment = data_alignment

def convert_to_excel(data):
    """Convert JSON data to Excel with formatted headers and aligned data."""
    try:
        data = validate_data(data)
        
        if not data:
            logger.warning("No valid data to export")
            return None

        wb = Workbook()
        ws = wb.active

        headers = [
            "JOB NO AND DATE", "IMPORTER", "SUPPLIER/ EXPORTER", "INVOICE NUMBER AND DATE",
            "INVOICE VALUE AND UNIT PRICE", "BL NUMBER AND DATE", "COMMODITY", "NET WEIGHT",
            "PORT", "ARRIVAL DATE", "FREE TIME", "DETENTION FROM", "SHIPPING LINE",
            "CONTAINER NUM & SIZE", "NUMBER OF CONTAINERS", "BE NUMBER AND DATE", "REMARKS", "DETAILED STATUS",
        ]
        ws.append(headers)

        # Apply column width settings
        set_column_widths(ws, headers)

        # Apply header styling
        style_header(ws, headers)

        start_data_row = 2  # Since headers are in row 1
        row_count = start_data_row  # Track data rows

        for row in data:
            containers = row.get('container_nos', [])
            job_no_date = f"{row.get('job_no', '')} | {row.get('job_date', '')} | {row.get('custom_house', '')} | {row.get('type_of_b_e', '')}"
            invoice_details = f"{row.get('invoice_number', '')} | {row.get('invoice_date', '')}"
            bl_details = f"{row.get('awb_bl_no', '')} | {row.get('awb_bl_date', '')}"
            remark_detail = (
                f"Discharge_Date: {row.get('discharge_date', '')} | "
                f"Arrival_Date: {row.get('assessment_date', '')} | "
                f"Duty_Paid_Date: {row.get('duty_paid_date', '')} | "
                f"DO_Validity_Upto_Job_Level: {row.get('do_validity_upto_job_level', '')}"
            )
            container_numbers = ", ".join(f"{c.get('container_number', '')} - {c.get('size', '')}" for c in row.get('container_nos', []))

            data_row = [
                job_no_date,
                row.get('importer', ''),
                row.get('supplier_exporter', ''),
                invoice_details,
                f"{row.get('inv_currency', '')} {row.get('invoice_value', '')} | {row.get('unit_price', '')}",
                bl_details,
                row.get('description', ''),
                row.get('job_net_weight', ''),
                f"POL: {row.get('loading_port', '')} POD: {row.get('port_of_reporting', '')}",
                format_container_dates(containers, 'arrival_date'),
                row.get('free_time', ''),
                format_container_dates(containers, 'detention_from'),
                row.get('shipping_line_airline', ''),
                container_numbers,
                row.get('no_of_container', ''),
                f"{row.get('be_no', '')} | {row.get('be_date', '')}",
                remark_detail,
                row.get('detailed_status', ''),
            ]
            ws.append(data_row)
            row_count += 1

        # Apply center alignment to data rows
        style_data(ws, start_data_row, row_count - 1, len(headers))

        output_filename = f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(output_filename)
        return output_filename

    except Exception as e:
        logger.error(f"Excel conversion failed: {e}")
        return None



# def convert_to_excel(data):
#     """Convert JSON data to Excel"""
#     try:
#         data = validate_data(data)
        

#         if not data:
#             logger.warning("No valid data to export")
#             return None

#         wb = Workbook()
#         ws = wb.active

#         headers = [
#             "JOB NO AND DATE", "IMPORTER", "SUPPLIER/ EXPORTER", "INVOICE NUMBER AND DATE",
#             "INVOICE VALUE AND UNIT PRICE", "BL NUMBER AND DATE", "COMMODITY", "NET WEIGHT",
#             "PORT", "ARRIVAL DATE", "FREE TIME", "DETENTION FROM", "SHIPPING LINE",
#             "CONTAINER NUM & SIZE", "NUMBER OF CONTAINERS", "BE NUMBER AND DATE", "REMARKS", "DETAILED STATUS",
#         ]
#         ws.append(headers)

#         for row in data:
#             containers = row.get('container_nos', [])
#             job_no_date = f"{row.get('job_no', '')} | {row.get('job_date', '')} | {row.get('custom_house', '')} | {row.get('type_of_b_e', '')}"
#             invoice_details = f"{row.get('invoice_number', '')} | {row.get('invoice_date', '')}"
#             bl_details = f"{row.get('awb_bl_no', '')} | {row.get('awb_bl_date', '')}"
#             remark_detail = (
#                         f"Discharge_Date: {row.get('discharge_date', '')} | "
#                         f"Arrival_Date: {row.get('assessment_date', '')} | "
#                         f"Duty_Paid_Date: {row.get('duty_paid_date', '')} | "
#                         f"DO_Validity_Upto_Job_Level: {row.get('do_validity_upto_job_level', '')}"
#                     )
#             container_numbers = ", ".join(f"{c.get('container_number', '')} - {c.get('size', '')}" for c in row.get('container_nos', []))

#             data_row = [
#                 job_no_date,
#                 row.get('importer', ''),
#                 row.get('supplier_exporter', ''),
#                 invoice_details,
#                 f"{row.get('inv_currency', '')} {row.get('invoice_value', '')} | {row.get('unit_price', '')}",
#                 bl_details,
#                 row.get('description', ''),
#                 row.get('job_net_weight', ''),
#                 f"POL: {row.get('loading_port', '')} POD: {row.get('port_of_reporting', '')}",
#                 format_container_dates(containers, 'arrival_date'),
#                 row.get('free_time', ''),
#                 format_container_dates(containers, 'detention_from'),
#                 row.get('shipping_line_airline', ''),
#                 container_numbers,
#                 row.get('no_of_container', ''),
#                 f"{row.get('be_no', '')} | {row.get('be_date', '')}",
#                 remark_detail,
#                 row.get('detailed_status', ''),
#             ]
#             ws.append(data_row)

#         output_filename = f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         wb.save(output_filename)
#         return output_filename

#     except Exception as e:
#         logger.error(f"Excel conversion failed: {e}")
#         return None



async def fetch_data():
    """Fetch API data and generate a report every 5 minutes"""
    global fetch_status
    while True:
        try:
            logger.info(f"Fetching data from {API_URL}")
            response = requests.get(API_URL)
            response.raise_for_status()
            data = response.json()

            output_filename = convert_to_excel(data)
            if output_filename:
                fetch_status["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                fetch_status["last_report"] = output_filename
                fetch_status["error"] = None
                logger.info(f"Excel report generated: {output_filename}")
            else:
                fetch_status["error"] = "Failed to generate report"

        except requests.RequestException as e:
            fetch_status["error"] = str(e)
            logger.error(f"API request failed: {e}")

        await asyncio.sleep(300)  # Wait for 5 minutes


@app.get("/status")
async def get_status():
    """Check the last data fetch status"""
    return fetch_status


@app.on_event("startup")
async def startup_event():
    """Start background data fetching when FastAPI starts"""
    background_task = asyncio.create_task(fetch_data())
    logger.info("Started background data fetch task")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
