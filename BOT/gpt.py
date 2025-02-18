from fastapi import FastAPI, HTTPException, BackgroundTasks
import pandas as pd
import uvicorn
import requests
import time
import os
import asyncio
from typing import Dict, Any
from threading import Thread, Event
from logging import getLogger, basicConfig, INFO
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)
basicConfig(level=INFO)
logger = getLogger(__name__)



app = FastAPI(title="Container Details API")

def validate_data(data):
    """Validate and extract the required API response data"""
    try:
        if isinstance(data, str):
            data = json.loads(data)

        if isinstance(data, dict) and "data" in data:
            data = data["data"]

        if not isinstance(data, list):
            raise ValueError("API response is not in expected format (list)")
        
        return data
    except Exception as e:
        logger.error(f"Data validation failed: {e}")
        raise

def format_date(date_str):
    """Formats the date field to avoid empty values"""
    return date_str if date_str else ""

def format_container_dates(containers, key):
    """Formats the container-related dates"""
    return ",\n".join(c.get(key, '') for c in containers)

def format_remarks(row):
    """Formats remarks field"""
    return row.get('remarks', '') if row.get('remarks') else ''

def set_column_widths(ws, headers):
    """Set column widths based on predefined values."""
    column_widths = {
        'JOB NO AND DATE': 40,
        'SUPPLIER/ EXPORTER': 40,
        'INVOICE NUMBER AND DATE': 25,
        'INVOICE VALUE AND UNIT PRICE': 35,
        'BL NUMBER AND DATE': 25,
        'COMMODITY': 50,
        'NET WEIGHT': 15,
        'PORT': 25,
        'ARRIVAL DATE': 15,
        'FREE TIME': 12,
        'DETENTION FROM': 15,
        'SHIPPING LINE': 40,
        'CONTAINER NUM & SIZE': 30,
        'WEIGHT EXCESS/SHORTAGE': 20,
        'NUMBER OF CONTAINERS': 15,
        'BE NUMBER AND DATE': 35,
        'REMARKS': 60,
        'DETAILED STATUS': 35 }

    for col_num, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = column_widths.get(header, 20)  # Default width 20

def style_header(ws, headers):
    """Apply styling to the header row (bold, centered, highlighted)."""
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light Yellow
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

def style_data(ws, start_row, end_row, num_columns):
    """Apply center alignment to all data cells."""
    data_alignment = Alignment(horizontal="center", vertical="center")

    for row in range(start_row, end_row + 1):
        for col in range(1, num_columns + 1):
            ws.cell(row=row, column=col).alignment = data_alignment

def convert_to_excel(data):
    """Convert JSON data to Excel with formatted headers and aligned data."""
    try:
        data = validate_data(data)
        if not data:
            logger.warning("No valid data to export")
            return None

        wb = Workbook()
        ws = wb.active
        headers = [
            "JOB NO AND DATE", "IMPORTER", "SUPPLIER/ EXPORTER", "INVOICE NUMBER AND DATE",
            "INVOICE VALUE AND UNIT PRICE", "BL NUMBER AND DATE", "COMMODITY", "NET WEIGHT",
            "PORT", "ARRIVAL DATE", "FREE TIME", "DETENTION FROM", "SHIPPING LINE",
            "CONTAINER NUM & SIZE", "NUMBER OF CONTAINERS", "BE NUMBER AND DATE", "REMARKS", "DETAILED STATUS",
        ]
        ws.append(headers)

        # Apply column width settings
        set_column_widths(ws, headers)

        # Apply header styling
        style_header(ws, headers)

        start_data_row = 2  # Since headers are in row 1
        row_count = start_data_row  # Track data rows

        for row in data:
            containers = row.get('container_nos', [])
            job_no_date = f"{row.get('job_no', '')} | {row.get('job_date', '')} | {row.get('custom_house', '')} | {row.get('type_of_b_e', '')}"
            invoice_details = f"{row.get('invoice_number', '')} | {row.get('invoice_date', '')}"
            bl_details = f"{row.get('awb_bl_no', '')} | {row.get('awb_bl_date', '')}"
            remark_detail = (
                f"Discharge_Date: {row.get('discharge_date', '')} | "
                f"Arrival_Date: {row.get('assessment_date', '')} | "
                f"Duty_Paid_Date: {row.get('duty_paid_date', '')} | "
                f"DO_Validity_Upto_Job_Level: {row.get('do_validity_upto_job_level', '')}"  )
            container_numbers = ", ".join(f"{c.get('container_number', '')} - {c.get('size', '')}" for c in row.get('container_nos', []))

            data_row = [
                job_no_date,
                row.get('importer', ''),
                row.get('supplier_exporter', ''),
                invoice_details,
                f"{row.get('inv_currency', '')} {row.get('invoice_value', '')} | {row.get('unit_price', '')}",
                bl_details,
                row.get('description', ''),
                row.get('job_net_weight', ''),
                f"POL: {row.get('loading_port', '')} POD: {row.get('port_of_reporting', '')}",
                format_container_dates(containers, 'arrival_date'),
                row.get('free_time', ''),
                format_container_dates(containers, 'detention_from'),
                row.get('shipping_line_airline', ''),
                container_numbers,
                row.get('no_of_container', ''),
                f"{row.get('be_no', '')} | {row.get('be_date', '')}",
                remark_detail,
                row.get('detailed_status', ''),
            ]
            ws.append(data_row)
            row_count += 1

        style_data(ws, start_data_row, row_count - 1, len(headers))

        output_filename = f"Namdeo.xlsx"
        wb.save(output_filename)
        return output_filename

    except Exception as e:
        logger.error(f"Excel conversion failed: {e}")
        return None

FILE_PATH = "output.xlsx"
def get_container_data(container_number: str):
    df = pd.read_excel(FILE_PATH, dtype=str).fillna("Not Available")
    df.columns = df.columns.astype(str).str.strip()
    
    container_column = "CONTAINER NUM & SIZE"
    if container_column not in df.columns:
        return {"error": f"Column '{container_column}' not found in the Excel file."}
    
    df[container_column] = df[container_column].str.upper()
    container_number = container_number.upper()
    filtered_df = df[df[container_column].str.contains(container_number, na=False, case=False)]
    
    if filtered_df.empty:
        return {"error": f"No data found for container number: {container_number}"}
    
    row = filtered_df.iloc[0]
    return row.to_dict()


def get_job_data(job_number: str):
    df = pd.read_excel(FILE_PATH, dtype=str).fillna("Not Available")
    df.columns = df.columns.astype(str).str.strip()
    
    job_column = "JOB NO AND DATE"
    if job_column not in df.columns:
        return {"error": f"Column '{job_column}' not found in the Excel file."}
    
    df[job_column] = df[job_column].str.upper()
    job_number = job_number.upper()
    filtered_df = df[df[job_column].str.contains(job_number, na=False, case=False)]
    
    if filtered_df.empty:
        return {"error": f"No data found for job number: {job_number}"}
    
    row = filtered_df.iloc[0]
    return row.to_dict()


async def fetch_data():
    """Fetch API data and generate a report every 5 minutes"""
    global fetch_status
    while True:
        try:
            API_URL = "http://43.205.59.159:9000/api/download-report/24-25/Pending"
            logger.info(f"Fetching data from {API_URL}")
            response = requests.get(API_URL)
            response.raise_for_status()
            data = response.json()

            output_filename = convert_to_excel(data)
            if output_filename:
                logger.info(f"Excel report generated: {output_filename}")
            else:
                fetch_status["error"] = "Failed to generate report"

        except requests.RequestException as e:
            fetch_status["error"] = str(e)
            logger.error(f"API request failed: {e}")

        await asyncio.sleep(300)  

@app.on_event("startup")
async def startup_event():
    """Start background data fetching when FastAPI starts"""
    background_task = asyncio.create_task(fetch_data())
    logger.info("Started background data fetch task")


@app.get("/container/{container_number}")
async def find_container_details(container_number: str):
    try:
        return get_container_data(container_number)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing container details: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while processing container details")


@app.get("/job/{job_number}")
async def find_job_details(job_number: str):
    try:
        return get_job_data(job_number)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing job details: {e}")
        raise HTTPException(status_code=500, detail="Internal server error while processing job details")


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8080)

