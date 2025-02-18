import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def json_to_excel(json_data, output_file="mm.xlsx"):
    columns = [
        'job_no', 'job_date', 'year', 'priorityJob', 'custom_house', 'importer',
        'supplier_exporter', 'invoice_number', 'invoice_date', 'assbl_value', 'awb_bl_no',
        'awb_bl_date', 'cif_amount', 'no_of_container', 'container_nos', 'cth_documents',
        'description', 'type_of_b_e', 'gross_weight', 'loading_port', 'origin_country',
        'port_of_reporting', 'shipping_line_airline', 'consignment_type', 'do_copies',
        'cth_no', 'total_duty', 'voyage_no', 'detailed_status', 'vessel_berthing',
        'vessel_flight', 'assessment_date', 'be_date', 'be_no', 'completed_operation_date',
        'inv_currency', 'job_owner', 'total_inv_value', 'status', 'shipping_line_attachment',
        'shipping_line_insurance', 'shipping_line_invoice_imgs', 'submissionQueries',
        'unit_1', 'utr', 'verified_checklist_upload', '__v', 'bill_document_sent_to_accounts',
        'containers_arrived_on_same_date', 'delivery_date', 'discharge_date', 'doPlanning',
        'do_completed', 'do_planning_date', 'do_revalidation', 'do_revalidation_date',
        'do_revalidation_upto_job_level', 'document_received_date', 'documentation_completed_date_time',
        'duty_paid_date', 'esanchit_completed_date_time', 'examinationPlanning', 'examination_planning_date',
        'free_time', 'gateway_igm_date', 'nfmims_date', 'nfmims_reg_no', 'obl_telex_bl',
        'out_of_charge', 'pims_date', 'pims_reg_no', 'remarks', 'sims_date', 'sims_reg_no',
        'submission_completed_date_time', 'type_of_Do', 'bill_date', 'bill_no', 'gateway_igm',
        'hss_name', 'igm_date', 'igm_no', 'no_of_pkgs', 'toi', 'unit', 'unit_price',
        'rail_out_date', 'do_validity_upto_job_level', 'do_processed', 'do_processed_date',
        'do_validity', 'other_invoices', 'other_invoices_date', 'payment_made', 'payment_made_date',
        'security_deposit', 'shipping_line_invoice', 'shipping_line_invoice_date', 'concor_gate_pass_date',
        'concor_gate_pass_validate_up_to', 'examination_date', 'pcv_date', 'fta_Benefit_date_time',
        'createdAt', 'updatedAt', 'custodian_gate_pass', 'custom_house', 'do_copies', 'do_documents',
        'do_queries', 'documentationQueries', 'documents', 'eSachitQueries', 'exrate',
        'gate_pass_copies', 'gross_weight', 'icd_cfs_invoice_img', 'importer', 'importerURL',
        'importer_address', 'inv_currency', 'is_free_time_updated', 'job_date', 'job_owner',
        'job_sticker_upload', 'loading_port', 'no_of_container', 'ooc_copies', 'origin_country',
        'other_invoices_img', 'port_of_reporting', 'processed_be_attachment' ]
    
    if isinstance(json_data, dict):
        json_data = [json_data]

    df = pd.DataFrame(json_data)
    existing_columns = [col for col in columns if col in df.columns]
    df = df[existing_columns]
    
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    for col_num, column_title in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 30
        for cell in ws[col_letter]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_file)
    print(f"Excel file '{output_file}' has been created successfully with formatted columns!")



import requests
API_URL = "http://43.205.59.159:9000/api/download-report/24-25/Pending"
response = requests.get(API_URL)
response.raise_for_status()
data = response.json()
output_file = json_to_excel(data)

